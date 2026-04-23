from __future__ import annotations

import os

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

FolderName = ["1-堺", "2-坂出", "3-八戸", "4-鹿島", "5-函館", 
              "6-青森", "7-釜石", "8-清水", "9-秋田", "10-金沢", 
              "11-苫小牧", "12-石巻", "13-小名浜", "14-神戸", "15-釧路", 
              "16-小樽", "17-呉", "18-新潟", "19-境港", "20-鹿児島"]
PortName = ["Sakai", "Sakaide", "Hachinohe", "Kashima", "Hakodate", 
            "Aomori", "Kamaishi", "Shimizu", "Akita", "Kanazawa", 
            "Tomakomai", "Ishinomaki", "Onahama", "Kobe", "Kushiro", 
            "Otaru", "Kure", "Nigata", "Sakaiminato", "Kagoshima"]


OUTPUT_COLUMNS = [
    "BCNSHP",
    "BOYSHP",
    "CATCAM",
    "CATLAM",
    "COLOUR",
    "COLPAT",
    "NOBJNAM",
    "OBJNAM",
    "latitude",
    "longitude",
]


KEY_MAP = {
    "BCNSHP": "BCNSHP",
    "BOYSHP": "BOYSHP",
    "CATCAM": "CATCAM",
    "CATLAM": "CATLAM",
    "COLOUR": "COLOUR",
    "COLPAT": "COLPAT",
    "NOBJNM": "NOBJNAM",
    "NOBJNAM": "NOBJNAM",
    "OBJNAM": "OBJNAM",
}


def normalize_value(value: Any) -> str:
    """
    CSVに書き込む値を文字列化する。
    None や 'null' は空文字にする。
    """
    if value is None:
        return ""

    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() == "null":
            return ""
        return stripped

    return str(value)


def to_float_if_possible(value: Any) -> Optional[float]:
    """
    数値に変換できるものは float にする。
    変換できなければ None。
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None

    return None


def detect_latitude_longitude_from_right(row_values: List[Any]) -> Tuple[str, str]:
    """
    行の右側から数値を2つ見つけ、
    100より大きい方を longitude、もう一方を latitude とする。
    """
    numeric_values = []

    for value in reversed(row_values):
        num = to_float_if_possible(value)
        if num is not None:
            numeric_values.append(num)
            if len(numeric_values) == 2:
                break

    if len(numeric_values) < 2:
        return "", ""

    a, b = numeric_values[0], numeric_values[1]

    if a > 100 and b <= 100:
        longitude, latitude = a, b
    elif b > 100 and a <= 100:
        longitude, latitude = b, a

    return normalize_value(latitude), normalize_value(longitude)


def parse_row(row_values: List[Any]) -> Dict[str, str]:
    """
    1行を解析して、必要な列だけ取り出す。
    行内の 'キー, 値, キー, 値, ...' 形式を想定。
    """
    row_data = {col: "" for col in OUTPUT_COLUMNS}

    for i in range(len(row_values) - 1):
        key = row_values[i]
        value = row_values[i + 1]

        if not isinstance(key, str):
            continue

        normalized_key = key.replace(":", "").strip().upper()

        if normalized_key in KEY_MAP:
            output_col = KEY_MAP[normalized_key]
            row_data[output_col] = normalize_value(value)

    latitude, longitude = detect_latitude_longitude_from_right(row_values)
    row_data["latitude"] = latitude
    row_data["longitude"] = longitude

    return row_data


def is_meaningful_record(record: Dict[str, str]) -> bool:
    """
    全項目空の行はCSVに入れない。
    """
    return any(value != "" for value in record.values())


def convert_excels_in_folder_to_csv(folder_path: str, csv_name: str) -> Path:
    """
    指定フォルダ内の複数のExcelファイルをすべて読み込み、
    連結して1つのCSVに保存する。

    読み込むシートは常に 'Sheet1'。
    """
    folder = Path(folder_path)

    if not csv_name.lower().endswith(".csv"):
        csv_name += ".csv"

    csv_dir = f"outputs/data/buoy"
    os.makedirs(csv_dir, exist_ok=True)

    excel_files = sorted(
        [
            p for p in folder.iterdir()
            if p.is_file()
            and not p.name.startswith("~$")
            and p.suffix.lower() in [".xlsx", ".xlsm"]
        ]
    )

    all_records = []
    skipped_files = []

    for excel_path in excel_files:
        try:
            wb = load_workbook(excel_path, data_only=True, read_only=True)

            ws = wb["Sheet1"]

            for row in ws.iter_rows(values_only=True):
                row_values = list(row)

                if all(v is None for v in row_values):
                    continue

                record = parse_row(row_values)

                if is_meaningful_record(record):
                    all_records.append(record)

        except Exception as e:
            skipped_files.append((excel_path.name, str(e)))

    csv_path = f"{csv_dir}/{csv_name}"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(all_records)

    print(f"CSV保存完了: {csv_path}")
    print(f"読み込んだExcelファイル数: {len(excel_files)}")
    print(f"出力行数: {len(all_records)}")

    if skipped_files:
        print("\nスキップしたファイル:")
        for file_name, reason in skipped_files:
            print(f"  - {file_name}: {reason}")

    return csv_path


if __name__ == "__main__":

    for folder_name, port_name in zip(FolderName, PortName):
        folder_path = f"raw_datas/buoy/{folder_name}"
        csv_name = f"{port_name}.csv"

        convert_excels_in_folder_to_csv(
            folder_path=folder_path,
            csv_name=csv_name,
        )